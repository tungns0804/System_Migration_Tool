"""Test sheet mo ta chi tiet tung method trong Excel (dot 9 — documents/03 muc 15).

Cover ma tran truong hop cua documents/14_Master_Prompt_v9.md:
cap 1-1, MISSING, EXTRA, 1 VB -> nhieu C# (related), body dai, tat tinh nang.
"""

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "tool"))

from openpyxl import load_workbook

from core.config import Config, set_config
from core.excel_exporter import export_excel, method_sheet_name
from core.models import (MethodInfo, MethodComparison, ScanResult,
                         STATUS_PASS, STATUS_MISSING, STATUS_EXTRA)


def _vb(name="GetPrice", body="Dim x As Integer\nReturn x", file="Form1.vb"):
    return MethodInfo(name=name, kind="function", return_type="Integer",
                      body=body, file=file, line=12,
                      signature=f"Function {name}() As Integer")


def _cs(name="GetPrice", body="int x = 0;\nreturn x;", file="Features\\Price\\PriceService.cs"):
    return MethodInfo(name=name, kind="typed", return_type="int",
                      body=body, file=file, line=34,
                      signature=f"int {name}()")


def _make_result() -> ScanResult:
    result = ScanResult()
    result.comparisons = [
        MethodComparison(name="GetPrice", vb=_vb(), cs=_cs(), status=STATUS_PASS,
                         score=95.0, similarity=0.91),
        MethodComparison(name="OldOnly", vb=_vb("OldOnly", file="Form2.vb"),
                         status=STATUS_MISSING,
                         notes=["Khong tim thay method tuong ung o he thong moi (C#)"]),
        MethodComparison(name="NewOnly", cs=_cs("NewOnly", file="Common\\Helper.cs"),
                         status=STATUS_EXTRA),
    ]
    return result


def _export(result, cfg=None) -> str:
    set_config(cfg or Config())
    out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    out.close()
    export_excel(result, out.name)
    return out.name


def _sheet_text(ws) -> str:
    return "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                     if c.value is not None)


class TestMethodSheets(unittest.TestCase):
    def tearDown(self):
        set_config(Config())

    def test_moi_dong_detail_mot_sheet(self):
        path = _export(_make_result())
        wb = load_workbook(path)
        self.assertEqual(wb.sheetnames, ["Summary", "Detail", "M001", "M002", "M003"])
        Path(path).unlink()

    def test_hyperlink_hai_chieu(self):
        path = _export(_make_result())
        wb = load_workbook(path)
        detail = wb["Detail"]
        # o ten method (cot B) tro dung sheet mo ta
        self.assertEqual(detail.cell(row=2, column=2).hyperlink.location, "M001!A1")
        self.assertEqual(detail.cell(row=4, column=2).hyperlink.location, "M003!A1")
        # sheet mo ta co link quay lai dung dong Detail
        m1 = wb["M001"]
        self.assertEqual(m1.cell(row=2, column=1).hyperlink.location, "Detail!B2")
        Path(path).unlink()

    def test_cap_1_1_du_2_muc_source(self):
        path = _export(_make_result())
        text = _sheet_text(load_workbook(path)["M001"])
        self.assertIn("SOURCE VB (HE THONG CU)", text)
        self.assertIn("Form1.vb : dong 12", text)
        self.assertIn("Function GetPrice() As Integer", text)
        self.assertIn("Dim x As Integer", text)          # body VB
        self.assertIn("SOURCE ASP.NET CORE (HE THONG MOI)", text)
        self.assertIn("PriceService.cs : dong 34", text)
        self.assertIn("int x = 0;", text)                # body C#
        Path(path).unlink()

    def test_missing_ghi_chua_implement(self):
        path = _export(_make_result())
        text = _sheet_text(load_workbook(path)["M002"])
        self.assertIn("CHUA IMPLEMENT tren ASP.NET Core", text)
        self.assertIn("Form2.vb : dong 12", text)        # nguon VB van day du
        self.assertIn("Khong tim thay method tuong ung", text)  # notes
        Path(path).unlink()

    def test_extra_ghi_khong_co_o_vb(self):
        path = _export(_make_result())
        text = _sheet_text(load_workbook(path)["M003"])
        self.assertIn("KHONG CO o source VB", text)
        self.assertIn("Common\\Helper.cs : dong 34", text)
        Path(path).unlink()

    def test_mot_vb_nhieu_cs_liet_ke_du(self):
        """1 VB ghep cap + 2 ban C# trung ten o folder khac -> lien ket cheo du."""
        result = ScanResult()
        result.comparisons = [
            MethodComparison(name="SaveOrder", status=STATUS_PASS,
                             vb=_vb("SaveOrder"),
                             cs=_cs("SaveOrder", file="Features\\Order\\Save.cs")),
            MethodComparison(name="SaveOrderAsync", status=STATUS_EXTRA,
                             cs=_cs("SaveOrderAsync", file="Features\\Draft\\SaveDraft.cs")),
            MethodComparison(name="SaveOrder", status=STATUS_EXTRA,
                             cs=_cs("SaveOrder", file="Legacy\\OrderUtil.cs")),
            MethodComparison(name="Unrelated", status=STATUS_EXTRA,
                             cs=_cs("Unrelated", file="X.cs")),
        ]
        path = _export(result)
        wb = load_workbook(path)
        text1 = _sheet_text(wb["M001"])
        self.assertIn("METHOD LIEN QUAN TRUNG TEN", text1)
        # liet ke DAY DU ca 2 ban trung ten (ke ca khop qua hau to Async) + file
        self.assertIn("SaveDraft.cs", text1)
        self.assertIn("OrderUtil.cs", text1)
        self.assertNotIn("Unrelated", text1)
        # chieu nguoc lai: sheet cua ban EXTRA cung tro ve cap chinh
        text3 = _sheet_text(wb["M003"])
        self.assertIn("M001", text3)
        # dong khong lien quan thi khong co muc lien quan
        self.assertNotIn("METHOD LIEN QUAN TRUNG TEN", _sheet_text(wb["M004"]))
        Path(path).unlink()

    def test_manh_tach_khai_bao_mapping_duoc_liet_ke(self):
        """Dot 11: A tach thanh B + C ten khac han — noi nhom qua related_names."""
        result = ScanResult()
        pair = MethodComparison(
            name="CloseCustomerAccount", status=STATUS_PASS,
            vb=_vb("CloseCustomerAccount"),
            cs=_cs("DeactivateCustomerAsync", file="Features\\CustomerMaster\\Deact.cs"))
        pair.related_names = ["cancelpendingorders"]
        piece = MethodComparison(
            name="CancelPendingOrdersAsync", status=STATUS_EXTRA,
            cs=_cs("CancelPendingOrdersAsync", file="Features\\OrderEntry\\Cancel.cs"))
        piece.related_names = ["closecustomeraccount"]
        result.comparisons = [pair, piece]
        path = _export(result)
        wb = load_workbook(path)
        t1 = _sheet_text(wb["M001"])
        self.assertIn("METHOD LIEN QUAN TRUNG TEN", t1)
        self.assertIn("Cancel.cs", t1)          # manh tach ten khac van duoc liet ke
        t2 = _sheet_text(wb["M002"])
        self.assertIn("Deact.cs", t2)           # chieu nguoc lai
        Path(path).unlink()

    def test_body_dai_bi_cat_kem_ghi_chu(self):
        result = ScanResult()
        long_body = "\n".join(f"Dim v{i} As Integer" for i in range(50))
        result.comparisons = [MethodComparison(
            name="Big", vb=_vb("Big", body=long_body), cs=_cs("Big"),
            status=STATUS_PASS)]
        cfg = Config({"excel": {"max_body_lines": 10}})
        path = _export(result, cfg)
        text = _sheet_text(load_workbook(path)["M001"])
        self.assertIn("Dim v9 As Integer", text)
        self.assertNotIn("Dim v10 As Integer", text)
        self.assertIn("con 40 dong", text)
        Path(path).unlink()

    def test_body_rong_in_ghi_chu(self):
        result = ScanResult()
        result.comparisons = [MethodComparison(
            name="Empty", vb=_vb("Empty", body=""), cs=_cs("Empty", body="  "),
            status=STATUS_PASS)]
        path = _export(result)
        text = _sheet_text(load_workbook(path)["M001"])
        self.assertEqual(text.count("(body rong)"), 2)
        Path(path).unlink()

    def test_tat_method_sheets(self):
        cfg = Config({"excel": {"method_sheets": False}})
        path = _export(_make_result(), cfg)
        wb = load_workbook(path)
        self.assertEqual(wb.sheetnames, ["Summary", "Detail"])
        # khong co hyperlink khi tat
        self.assertIsNone(wb["Detail"].cell(row=2, column=2).hyperlink)
        Path(path).unlink()

    def test_detail_giu_nguyen_cot_va_dropdown(self):
        """Khong degrade: bo cot A->S (dot 12 chen 'AI de xuat giai phap' canh
        Status AI), dropdown DEV, auto-filter giu nguyen."""
        path = _export(_make_result())
        ws = load_workbook(path)["Detail"]
        self.assertEqual(ws.cell(row=1, column=15).value, "Notes")
        self.assertEqual(ws.cell(row=1, column=16).value, "Nội dung AI đánh giá")
        self.assertEqual(ws.cell(row=1, column=17).value, "Status AI đánh giá")
        self.assertEqual(ws.cell(row=1, column=18).value, "AI đề xuất giải pháp")
        self.assertEqual(ws.cell(row=1, column=19).value, "Status DEV đánh giá")
        self.assertEqual(ws.auto_filter.ref, "A1:S4")
        self.assertEqual(len(list(ws.data_validations.dataValidation)), 1)
        Path(path).unlink()


class TestTimestampExport(unittest.TestCase):
    """Dot 9b: token {ts} trong duong dan export -> timestamp, tranh trung ten."""

    def test_token_duoc_thay_bang_timestamp(self):
        from datetime import datetime
        from core.report_io import apply_timestamp
        when = datetime(2026, 7, 16, 18, 45, 7)
        self.assertEqual(apply_timestamp("report_{ts}.xlsx", when),
                         "report_20260716_184507.xlsx")
        self.assertEqual(apply_timestamp(r"out\{ts}\bao_cao_{ts}.json", when),
                         r"out\20260716_184507\bao_cao_20260716_184507.json")

    def test_khong_token_giu_nguyen(self):
        from core.report_io import apply_timestamp
        self.assertEqual(apply_timestamp("samples/migration_report_pcrs.xlsx"),
                         "samples/migration_report_pcrs.xlsx")
        self.assertEqual(apply_timestamp(""), "")


if __name__ == "__main__":
    unittest.main()
