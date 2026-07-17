"""Test gate danh gia AI (dot 6 — documents/03 muc 13).

Mock hoan toan qua call_fn — KHONG goi Claude API that, khong can mang,
khong can cai package anthropic.
"""

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "tool"))

from core.config import Config, set_config
from core.excel_exporter import export_excel
from core.llm_reviewer import (review_result, cache_key, build_payload,
                               load_cache, PROMPT_VERSION,
                               resolve_provider, effective_model,
                               _parse_google_response, _google_request_body,
                               _google_retry_delay,
                               ANTHROPIC_DEFAULT_MODEL, GOOGLE_DEFAULT_MODEL)
from core.models import (MethodInfo, MethodComparison, ScanResult,
                         STATUS_PASS, STATUS_MISSING, STATUS_EXTRA,
                         AI_PASS, AI_WARNING, AI_NOT_RUN)
from core.report_io import result_to_dict


def _vb(name="GetPrice", body="Return 1"):
    return MethodInfo(name=name, kind="function", return_type="Integer",
                      body=body, file="Form1.vb", signature=f"Function {name}()")


def _cs(name="GetPrice", body="return 1;"):
    return MethodInfo(name=name, kind="typed", return_type="int",
                      body=body, file="Svc.cs", signature=f"int {name}()")


def _make_result(tmp_cache: str) -> ScanResult:
    """1 cap du 2 ben + 1 MISSING + 1 EXTRA."""
    result = ScanResult()
    pair = MethodComparison(name="GetPrice", vb=_vb(), cs=_cs(), status=STATUS_PASS)
    missing = MethodComparison(name="OldOnly", vb=_vb("OldOnly", "Return 2"),
                               status=STATUS_MISSING)
    extra = MethodComparison(name="NewOnly", cs=_cs("NewOnly", "return 3;"),
                             status=STATUS_EXTRA)
    result.comparisons = [pair, missing, extra]
    return result


def _cfg(tmp_cache: str, enabled=True) -> Config:
    return Config({"llm": {"enabled": enabled, "cache_file": tmp_cache}})


def _tmp_cache() -> str:
    f = tempfile.NamedTemporaryFile("w", suffix=".json", delete=False)
    f.close()
    Path(f.name).unlink()  # chi lay duong dan — file chua ton tai
    return f.name


class TestDefaults(unittest.TestCase):
    """Mac dinh llm tat — hanh vi cu giu nguyen 100%."""

    def test_llm_tat_mac_dinh(self):
        cfg = Config()
        self.assertFalse(cfg.llm_enabled)
        self.assertEqual(cfg.llm("model"), "claude-sonnet-5")
        self.assertEqual(cfg.llm("api_key"), "")

    def test_field_ai_mac_dinh_rong(self):
        comp = MethodComparison(name="X")
        self.assertEqual(comp.ai_status, "")
        self.assertEqual(comp.ai_comment, "")

    def test_json_ai_null_khi_chua_chay(self):
        result = _make_result("")
        data = result_to_dict(result)
        for m in data["methods"]:
            self.assertIsNone(m["ai_status"])
            self.assertIsNone(m["ai_comment"])

    def test_key_ghi_chu_gach_duoi_duoc_bo_qua(self):
        # config.sample.json co "_huong_dan"/"_llm_ghi_chu" — khong duoc crash
        cfg = Config({"_huong_dan": "abc", "llm": {"enabled": True}})
        self.assertTrue(cfg.llm_enabled)


class TestReview(unittest.TestCase):
    def test_cham_du_3_loai_method(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        calls = []

        def fake(payload):
            calls.append(payload)
            return "PASS", "OK het"

        stats = review_result(result, _cfg(cache), call_fn=fake)
        self.assertEqual(stats["reviewed"], 3)
        self.assertEqual(len(calls), 3)
        for comp in result.comparisons:
            self.assertEqual(comp.ai_status, AI_PASS)
            self.assertEqual(comp.ai_comment, "OK het")
        # payload MISSING/EXTRA phai neu ro chi ton tai 1 phia
        self.assertIn("MISSING", calls[1])
        self.assertIn("EXTRA", calls[2])

    def test_status_la_bay_ve_warning(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        stats = review_result(result, _cfg(cache),
                              call_fn=lambda p: ("FAIL", "nghi ngo"))
        self.assertEqual(stats["reviewed"], 3)
        for comp in result.comparisons:
            self.assertEqual(comp.ai_status, AI_WARNING)
            self.assertIn("FAIL", comp.ai_comment)

    def test_khong_dung_den_diem_va_status_cu(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        before = [(c.status, c.score, list(c.notes)) for c in result.comparisons]
        review_result(result, _cfg(cache), call_fn=lambda p: ("WARNING", "x"))
        after = [(c.status, c.score, list(c.notes)) for c in result.comparisons]
        self.assertEqual(before, after)

    def test_cache_hit_khong_goi_lai(self):
        cache = _tmp_cache()
        cfg = _cfg(cache)
        result1 = _make_result(cache)
        calls = []
        fake = lambda p: (calls.append(p) or ("PASS", "lan dau"))
        review_result(result1, cfg, call_fn=fake)
        self.assertEqual(len(calls), 3)

        # Lan 2 cung noi dung -> toan bo tu cache, khong goi call_fn
        result2 = _make_result(cache)
        stats = review_result(result2, cfg,
                              call_fn=lambda p: self.fail("khong duoc goi API khi cache hit"))
        self.assertEqual(stats["cached"], 3)
        self.assertEqual(result2.comparisons[0].ai_comment, "lan dau")

        # Doi body -> hash khac -> goi lai dung 1 method do
        result3 = _make_result(cache)
        result3.comparisons[0].cs.body = "return 999;"
        calls3 = []
        review_result(result3, cfg, call_fn=lambda p: (calls3.append(p) or ("WARNING", "moi")))
        self.assertEqual(len(calls3), 1)
        self.assertEqual(result3.comparisons[0].ai_status, AI_WARNING)

    def test_cache_key_phu_thuoc_model_va_prompt_version(self):
        comp = MethodComparison(name="X", vb=_vb(), cs=_cs())
        k1 = cache_key(comp, "claude-sonnet-5")
        k2 = cache_key(comp, "model-khac")
        self.assertNotEqual(k1, k2)
        self.assertEqual(PROMPT_VERSION, "v2")  # doi prompt phai tang version

    def test_loi_le_te_khong_chan_method_sau(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        seq = iter([RuntimeError("timeout"), ("PASS", "ok"), ("PASS", "ok")])

        def flaky(payload):
            item = next(seq)
            if isinstance(item, Exception):
                raise item
            return item

        stats = review_result(result, _cfg(cache), call_fn=flaky)
        self.assertEqual(result.comparisons[0].ai_status, AI_NOT_RUN)
        self.assertIn("timeout", result.comparisons[0].ai_comment)
        self.assertEqual(result.comparisons[1].ai_status, AI_PASS)
        self.assertEqual(stats["reviewed"], 2)
        self.assertEqual(stats["not_run"], 1)

    def test_thieu_key_tat_ca_not_run(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        cfg = Config({"llm": {"enabled": True, "cache_file": cache,
                              "api_key_env": "BIEN_MOI_TRUONG_CHAC_CHAN_KHONG_TON_TAI"}})
        stats = review_result(result, cfg)  # call_fn=None -> tu tao caller that
        self.assertIsNotNone(stats["error"])
        self.assertEqual(stats["not_run"], 3)
        for comp in result.comparisons:
            self.assertEqual(comp.ai_status, AI_NOT_RUN)
            self.assertIn("Khong goi duoc AI API", comp.ai_comment)

    def test_het_credit_dung_goi_ngay(self):
        """Loi 'credit balance too low' la loi chi mang — chi goi API 1 lan."""
        cache = _tmp_cache()
        result = _make_result(cache)
        calls = []

        def no_credit(payload):
            calls.append(payload)
            raise RuntimeError(
                "Error code: 400 - Your credit balance is too low to access the Anthropic API.")

        stats = review_result(result, _cfg(cache), call_fn=no_credit)
        self.assertEqual(len(calls), 1)          # khong goi lai 42 lan vo ich
        self.assertEqual(stats["not_run"], 3)
        self.assertIn("het credit", stats["error"])
        for comp in result.comparisons:
            self.assertEqual(comp.ai_status, AI_NOT_RUN)

    def test_review_khong_bao_gio_raise(self):
        cache = _tmp_cache()
        result = _make_result(cache)

        def always_boom(payload):
            raise ConnectionError("mat mang")

        stats = review_result(result, _cfg(cache), call_fn=always_boom)
        self.assertEqual(stats["not_run"], 3)

    def test_thong_ke_count_ai(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        seq = iter([("PASS", "a"), ("WARNING", "b"), RuntimeError("x")])

        def fake(payload):
            item = next(seq)
            if isinstance(item, Exception):
                raise item
            return item

        review_result(result, _cfg(cache), call_fn=fake)
        counts = result.count_ai()
        self.assertEqual(counts[AI_PASS], 1)
        self.assertEqual(counts[AI_WARNING], 1)
        self.assertEqual(counts["not_run"], 1)
        self.assertTrue(result.ai_reviewed())

    def test_progress_bao_du_moi_method(self):
        """Dot 8: callback progress duoc goi cho MOI method (ke ca cache hit va loi)."""
        cache = _tmp_cache()
        cfg = _cfg(cache)
        result = _make_result(cache)
        ticks = []
        review_result(result, cfg, call_fn=lambda p: ("PASS", "ok"),
                      progress=lambda i, total, cached: ticks.append((i, total, cached)))
        self.assertEqual(ticks, [(1, 3, False), (2, 3, False), (3, 3, False)])

        # lan 2: toan bo cache hit -> van bao du 3 tick, cached=True
        result2 = _make_result(cache)
        ticks2 = []
        review_result(result2, cfg, call_fn=lambda p: ("PASS", "ok"),
                      progress=lambda i, total, cached: ticks2.append(cached))
        self.assertEqual(ticks2, [True, True, True])

        # loi chi mang: cac method con lai van duoc bao tien do
        result3 = _make_result(cache)
        result3.comparisons[0].vb.body = "Return 99"  # 1 cache miss dau tien
        ticks3 = []

        def boom(payload):
            raise RuntimeError("Your credit balance is too low")

        review_result(result3, cfg, call_fn=boom,
                      progress=lambda i, total, cached: ticks3.append(i))
        self.assertEqual(ticks3, [1, 2, 3])
        Path(cache).unlink()

    def test_payload_chua_du_thong_tin(self):
        comp = MethodComparison(name="GetPrice", vb=_vb(), cs=_cs(),
                                status=STATUS_PASS, notes=["C5: ghi chu tool"])
        payload = build_payload(comp)
        self.assertIn("GetPrice", payload)
        self.assertIn("Return 1", payload)      # body VB
        self.assertIn("return 1;", payload)     # body C#
        self.assertIn("C5: ghi chu tool", payload)


def _dict_reply(status="WARNING", **overrides):
    """Ket qua call_fn kieu moi (dot 12) — dict day du 5 truong."""
    data = {
        "status": status,
        "comment": "cần review chỗ làm tròn",
        "suggestion_overview": "Đổi (int) sang Math.Round để giữ banker's rounding.",
        "suggestion_detail": "1. CInt của VB làm tròn banker's rounding.\n"
                             "2. Sửa dòng return dùng Math.Round.",
        "suggestion_code": "int GetPrice()\n{\n"
                           "    return (int)Math.Round(x, MidpointRounding.ToEven); "
                           "// FIX: giu banker's rounding nhu CInt\n}",
    }
    data.update(overrides)
    return data


class TestSuggestion(unittest.TestCase):
    """Dot 12: AI de xuat giai phap cho cho khong PASS."""

    def test_warning_gan_du_3_truong(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        stats = review_result(result, _cfg(cache), call_fn=lambda p: _dict_reply())
        self.assertEqual(stats["reviewed"], 3)
        comp = result.comparisons[0]
        self.assertEqual(comp.ai_status, AI_WARNING)
        self.assertIn("Math.Round", comp.ai_suggestion)
        self.assertIn("1.", comp.ai_suggestion_detail)
        self.assertIn("// FIX", comp.ai_suggestion_code)
        Path(cache).unlink()

    def test_pass_ep_suggestion_ve_rong(self):
        """Model tra PASS nhung van dinh kem suggestion -> phai bi xoa sach."""
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache), call_fn=lambda p: _dict_reply(status="PASS"))
        for comp in result.comparisons:
            self.assertEqual(comp.ai_status, AI_PASS)
            self.assertEqual(comp.ai_suggestion, "")
            self.assertEqual(comp.ai_suggestion_detail, "")
            self.assertEqual(comp.ai_suggestion_code, "")
        Path(cache).unlink()

    def test_call_fn_tuple_cu_van_chay(self):
        """Tuong thich nguoc: call_fn tra (status, comment) -> suggestion rong."""
        cache = _tmp_cache()
        result = _make_result(cache)
        stats = review_result(result, _cfg(cache), call_fn=lambda p: ("WARNING", "x"))
        self.assertEqual(stats["reviewed"], 3)
        comp = result.comparisons[0]
        self.assertEqual(comp.ai_status, AI_WARNING)
        self.assertEqual(comp.ai_comment, "x")
        self.assertEqual(comp.ai_suggestion, "")
        self.assertEqual(comp.ai_suggestion_code, "")
        Path(cache).unlink()

    def test_cache_giu_suggestion(self):
        cache = _tmp_cache()
        cfg = _cfg(cache)
        review_result(_make_result(cache), cfg, call_fn=lambda p: _dict_reply())
        result2 = _make_result(cache)
        stats = review_result(result2, cfg,
                              call_fn=lambda p: self.fail("phai lay tu cache"))
        self.assertEqual(stats["cached"], 3)
        comp = result2.comparisons[0]
        self.assertIn("Math.Round", comp.ai_suggestion)
        self.assertIn("banker", comp.ai_suggestion_detail)
        self.assertIn("// FIX", comp.ai_suggestion_code)
        Path(cache).unlink()

    def test_bo_code_fence_markdown(self):
        """Phong thu: model boc code trong ``` du prompt cam -> go bo."""
        from core.llm_reviewer import _strip_code_fence
        fenced = "```csharp\nint A() { return 1; } // FIX: x\n```"
        self.assertEqual(_strip_code_fence(fenced),
                         "int A() { return 1; } // FIX: x")
        self.assertEqual(_strip_code_fence("int B();"), "int B();")

    def test_json_co_truong_suggestion(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache), call_fn=lambda p: _dict_reply())
        m = result_to_dict(result)["methods"][0]
        self.assertIn("Math.Round", m["ai_suggestion"])
        self.assertIn("banker", m["ai_suggestion_detail"])
        self.assertIn("// FIX", m["ai_suggestion_code"])
        Path(cache).unlink()

    def test_json_suggestion_null_khi_chua_chay(self):
        result = _make_result("")
        for m in result_to_dict(result)["methods"]:
            self.assertIsNone(m["ai_suggestion"])
            self.assertIsNone(m["ai_suggestion_detail"])
            self.assertIsNone(m["ai_suggestion_code"])


class TestExcelSuggestion(unittest.TestCase):
    """Dot 12: cot 'AI de xuat giai phap' (Detail chi co overview) + muc
    de xuat highlight trong sheet mo ta Mxxx."""

    @staticmethod
    def _sheet_text(ws):
        return "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                         if c.value is not None)

    def _export(self, result):
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        out.close()
        export_excel(result, out.name)
        return out.name

    def test_cot_overview_va_sheet_chi_tiet(self):
        from openpyxl import load_workbook
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache), call_fn=lambda p: _dict_reply())
        path = self._export(result)
        wb = load_workbook(path)

        # Detail: cot R (18) ngay ben phai 'Status AI danh gia' — CHI overview
        ws = wb["Detail"]
        self.assertEqual(ws.cell(row=1, column=18).value, "AI đề xuất giải pháp")
        sugg = ws.cell(row=2, column=18).value
        self.assertIn("Math.Round", sugg)         # tom tat huong sua
        self.assertIn("sheet M001", sugg)         # chi dan sang sheet chi tiet
        self.assertNotIn("// FIX", sugg)          # khong dan code vao Detail
        # o co de xuat duoc to nen vang de nhan biet
        self.assertTrue(str(ws.cell(row=2, column=18).fill.fgColor.rgb).endswith("FFF2CC"))

        # Sheet Mxxx: du tom tat + giai thich chi tiet + code de xuat
        m1 = wb["M001"]
        text = self._sheet_text(m1)
        self.assertIn("AI DE XUAT GIAI PHAP", text)
        self.assertIn("Tom tat huong sua", text)
        self.assertIn("Giai thich chi tiet", text)
        self.assertIn("banker's rounding", text)
        self.assertIn("Code de xuat", text)
        self.assertIn("Math.Round", text)

        # dong code AI sua ("// FIX") highlight dam + bold; dong code thuong nen vang nhat
        cells = [c for row in m1.iter_rows() for c in row
                 if isinstance(c.value, str)]
        fix_cell = next(c for c in cells if "// FIX" in c.value)
        self.assertTrue(str(fix_cell.fill.fgColor.rgb).endswith("FFE08A"))
        self.assertTrue(fix_cell.font.bold)
        normal_code = next(c for c in cells if c.value.strip() == "int GetPrice()")
        self.assertTrue(str(normal_code.fill.fgColor.rgb).endswith("FFF2CC"))
        Path(path).unlink()
        Path(cache).unlink()

    def test_pass_khong_co_muc_de_xuat(self):
        from openpyxl import load_workbook
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache), call_fn=lambda p: _dict_reply(status="PASS"))
        path = self._export(result)
        wb = load_workbook(path)
        self.assertIsNone(wb["Detail"].cell(row=2, column=18).value)
        self.assertNotIn("AI DE XUAT GIAI PHAP", self._sheet_text(wb["M001"]))
        Path(path).unlink()
        Path(cache).unlink()

    def test_warning_khong_code_van_co_giai_thich(self):
        """AI de xuat khong kem code (chi can xac nhan tay) -> van hien giai thich."""
        from openpyxl import load_workbook
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache),
                      call_fn=lambda p: _dict_reply(suggestion_code=""))
        path = self._export(result)
        text = self._sheet_text(load_workbook(path)["M001"])
        self.assertIn("AI DE XUAT GIAI PHAP", text)
        self.assertIn("AI khong de xuat code", text)
        Path(path).unlink()
        Path(cache).unlink()

    def test_tat_method_sheets_khong_tro_sheet(self):
        """method_sheets=false -> o overview khong duoc tro den sheet Mxxx."""
        from openpyxl import load_workbook
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache), call_fn=lambda p: _dict_reply())
        set_config(Config({"excel": {"method_sheets": False}}))
        try:
            path = self._export(result)
        finally:
            set_config(Config())
        ws = load_workbook(path)["Detail"]
        sugg = ws.cell(row=2, column=18).value
        self.assertIn("Math.Round", sugg)
        self.assertNotIn("sheet M001", sugg)
        Path(path).unlink()
        Path(cache).unlink()


class TestProvider(unittest.TestCase):
    """Dot 8: nhan dien provider theo key + model thuc dung."""

    def _cfg_key(self, key, provider="auto", model="claude-sonnet-5"):
        return Config({"llm": {"api_key": key, "provider": provider, "model": model}})

    def test_auto_theo_dang_key(self):
        self.assertEqual(resolve_provider(self._cfg_key("AIzaXXXX")), "google")
        self.assertEqual(resolve_provider(self._cfg_key("sk-ant-xxx")), "anthropic")
        self.assertEqual(resolve_provider(self._cfg_key("")), "anthropic")

    def test_ep_provider_thang(self):
        self.assertEqual(resolve_provider(self._cfg_key("AIzaXXXX", provider="anthropic")),
                         "anthropic")
        self.assertEqual(resolve_provider(self._cfg_key("sk-ant-x", provider="google")),
                         "google")

    def test_model_thuc_dung_theo_provider(self):
        cfg = self._cfg_key("AIzaXXXX")  # model mac dinh claude-sonnet-5
        self.assertEqual(effective_model(cfg, "google"), GOOGLE_DEFAULT_MODEL)
        self.assertEqual(effective_model(cfg, "anthropic"), "claude-sonnet-5")
        cfg2 = self._cfg_key("AIzaXXXX", model="gemini-2.5-pro")
        self.assertEqual(effective_model(cfg2, "google"), "gemini-2.5-pro")
        self.assertEqual(effective_model(cfg2, "anthropic"), ANTHROPIC_DEFAULT_MODEL)

    def test_cache_key_dung_model_thuc_dung(self):
        # key Google + model claude trong config -> cache phai theo gemini
        cache = _tmp_cache()
        cfg = Config({"llm": {"enabled": True, "cache_file": cache,
                              "api_key": "AIzaXXXX"}})
        result = _make_result(cache)
        review_result(result, cfg, call_fn=lambda p: ("PASS", "ok"))
        data = load_cache(Path(cache))
        expected = cache_key(result.comparisons[0], GOOGLE_DEFAULT_MODEL)
        self.assertIn(expected, data)
        Path(cache).unlink()

    def test_provider_la_bao_loi_ro(self):
        cache = _tmp_cache()
        cfg = Config({"llm": {"enabled": True, "cache_file": cache,
                              "api_key": "AIzaXXXX", "provider": "openai"}})
        result = _make_result(cache)
        stats = review_result(result, cfg)
        self.assertIn("khong ho tro", stats["error"])

    def test_het_credit_google_dung_goi_ngay(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        calls = []

        def depleted(payload):
            calls.append(payload)
            raise RuntimeError(
                "HTTP 429: Your prepayment credits are depleted. RESOURCE_EXHAUSTED")

        stats = review_result(result, _cfg(cache), call_fn=depleted)
        self.assertEqual(len(calls), 1)          # khong retry/goi lai vo ich
        self.assertIn("aistudio.google.com", stats["error"])
        for comp in result.comparisons:
            self.assertEqual(comp.ai_status, AI_NOT_RUN)

    def test_sai_key_google_dung_goi_ngay(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        calls = []

        def bad_key(payload):
            calls.append(payload)
            raise RuntimeError("HTTP 400: API key not valid. Please pass a valid API key.")

        stats = review_result(result, _cfg(cache), call_fn=bad_key)
        self.assertEqual(len(calls), 1)
        self.assertIn("aistudio.google.com", stats["error"])


class TestGoogleCaller(unittest.TestCase):
    """Dot 8: dung request/parse response Gemini (khong goi mang)."""

    def test_request_body_du_thanh_phan(self):
        body = _google_request_body("noi dung method", "gemini-2.5-flash", 4096)
        self.assertIn("system_instruction", body)
        self.assertEqual(body["contents"][0]["parts"][0]["text"], "noi dung method")
        gen = body["generationConfig"]
        self.assertEqual(gen["responseMimeType"], "application/json")
        self.assertEqual(gen["responseSchema"]["properties"]["status"]["enum"],
                         ["PASS", "WARNING"])
        self.assertEqual(gen["maxOutputTokens"], 4096)

    def test_thinking_tat_cho_flash_giu_cho_pro(self):
        # Flash: tat thinking de nhanh + do ton quota; Pro khong cho tat
        flash = _google_request_body("x", "gemini-2.5-flash", 1024)
        self.assertEqual(flash["generationConfig"]["thinkingConfig"],
                         {"thinkingBudget": 0})
        pro = _google_request_body("x", "gemini-2.5-pro", 1024)
        self.assertNotIn("thinkingConfig", pro["generationConfig"])

    def test_parse_response_ok(self):
        data = {"candidates": [{"content": {"parts": [
            {"text": '{"status": "WARNING", "comment": "cần xem lại làm tròn", '
                     '"suggestion_overview": "Dùng Math.Round."}'}]}}]}
        obj = _parse_google_response(data)  # dot 12: tra dict day du
        self.assertEqual(obj["status"], "WARNING")
        self.assertEqual(obj["comment"], "cần xem lại làm tròn")
        self.assertEqual(obj["suggestion_overview"], "Dùng Math.Round.")

    def test_parse_response_loi_va_block(self):
        with self.assertRaises(RuntimeError):
            _parse_google_response({"error": {"message": "API key not valid"}})
        with self.assertRaises(RuntimeError):
            _parse_google_response({"promptFeedback": {"blockReason": "SAFETY"}})

    def test_retry_delay_doc_tu_response(self):
        detail = '{"error": {"details": [{"retryDelay": "18s"}]}}'
        self.assertEqual(_google_retry_delay(detail, 0), 19.0)
        self.assertEqual(_google_retry_delay("khong co", 1), 30.0)


class TestExcel(unittest.TestCase):
    def test_hai_cot_ai_trong_excel(self):
        from openpyxl import load_workbook
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache),
                      call_fn=lambda p: ("WARNING", "cần review chỗ làm tròn"))
        result.comparisons[2].ai_status = ""  # gia lap 1 dong chua chay AI

        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        out.close()
        export_excel(result, out.name)
        ws = load_workbook(out.name)["Detail"]
        self.assertEqual(ws.cell(row=1, column=16).value, "Nội dung AI đánh giá")
        self.assertEqual(ws.cell(row=1, column=17).value, "Status AI đánh giá")
        self.assertEqual(ws.cell(row=2, column=16).value, "cần review chỗ làm tròn")
        self.assertEqual(ws.cell(row=2, column=17).value, AI_WARNING)
        # dong chua chay AI -> fallback "AI chua thuc hien danh gia"
        self.assertEqual(ws.cell(row=4, column=17).value, AI_NOT_RUN)
        self.assertEqual(ws.auto_filter.ref, "A1:S4")
        Path(out.name).unlink()

    def test_cot_dev_dropdown(self):
        """Dot 7 (dot 12 doi R->S): cot 'Status DEV danh gia' — dropdown 5 status, o de trong."""
        from openpyxl import load_workbook
        result = _make_result("")
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        out.close()
        export_excel(result, out.name)
        ws = load_workbook(out.name)["Detail"]
        self.assertEqual(ws.cell(row=1, column=19).value, "Status DEV đánh giá")
        for r in range(2, 5):
            self.assertIsNone(ws.cell(row=r, column=19).value)  # de trong cho DEV
        dvs = list(ws.data_validations.dataValidation)
        self.assertEqual(len(dvs), 1)
        self.assertEqual(dvs[0].type, "list")
        self.assertEqual(dvs[0].formula1, '"PASS,WARNING,FAIL,MISSING,EXTRA"')
        self.assertIn("S2:S4", str(dvs[0].sqref))
        # co conditional formatting to mau theo 5 status
        self.assertTrue(any("S2:S4" in str(rng) for rng in ws.conditional_formatting))
        Path(out.name).unlink()

    def test_excel_khi_llm_tat_van_co_cot_fallback(self):
        from openpyxl import load_workbook
        result = _make_result("")
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        out.close()
        export_excel(result, out.name)
        ws = load_workbook(out.name)["Detail"]
        for r in range(2, 5):
            self.assertEqual(ws.cell(row=r, column=17).value, AI_NOT_RUN)
        Path(out.name).unlink()


class TestCacheFile(unittest.TestCase):
    def test_cache_ghi_va_doc_lai_duoc(self):
        cache = _tmp_cache()
        result = _make_result(cache)
        review_result(result, _cfg(cache), call_fn=lambda p: ("PASS", "ok"))
        data = load_cache(Path(cache))
        self.assertEqual(len(data), 3)
        for v in data.values():
            self.assertEqual(v["status"], AI_PASS)
        Path(cache).unlink()

    def test_cache_hong_khong_crash(self):
        cache = _tmp_cache()
        Path(cache).write_text("{khong phai json", encoding="utf-8")
        result = _make_result(cache)
        stats = review_result(result, _cfg(cache), call_fn=lambda p: ("PASS", "ok"))
        self.assertEqual(stats["reviewed"], 3)
        Path(cache).unlink()


if __name__ == "__main__":
    unittest.main()
