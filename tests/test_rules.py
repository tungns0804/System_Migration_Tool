# -*- coding: utf-8 -*-
"""Unit test cho rule-based check dot 4 (rules/conversion_rules.json + rule_checker)."""

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "tool"))

from core.comparator import scan_folders
from core.config import Config, set_config
from core.models import MethodInfo, CriterionResult
from core.rule_checker import (Rules, load_rules, find_default_rules,
                               check_vb_side, check_cs_side, check_pair_rules,
                               validator_split_note, set_rules)

_RULES_FILE = Path(__file__).parent.parent / "rules" / "conversion_rules.json"


def _rules() -> Rules:
    return load_rules(str(_RULES_FILE))


def _mk(name, body, return_type="void"):
    return MethodInfo(name=name, kind="method", return_type=return_type, body=body)


class TestRuleFile(unittest.TestCase):
    def test_file_ton_tai_va_doc_duoc(self):
        self.assertTrue(_RULES_FILE.is_file(), "rules/conversion_rules.json phai co trong repo")
        rules = _rules()
        self.assertFalse(rules.empty)

    def test_du_bo_rule_theo_documents_09(self):
        rules = _rules()
        vb_ids = {r[0] for r in rules.vb_patterns}
        cs_ids = {r[0] for r in rules.cs_patterns}
        self.assertEqual(vb_ids, {"VB-CINT", "VB-INTDIV", "VB-IIF", "VB-ONERROR",
                                  "VB-ROWS0", "VB-EMPTYSTR", "VB-ANDOR"})
        self.assertEqual(cs_ids, {"CS-WAITRESULT", "CS-FIREFORGET"})
        for name in ("jp_message", "self_exclusion", "validator_split", "rownum_order"):
            self.assertTrue(rules.pair_cfg(name), f"pair rule '{name}' phai enabled")

    def test_tu_tim_thay_file_mac_dinh(self):
        self.assertTrue(find_default_rules().endswith("conversion_rules.json"))


class TestRulesExcelExport(unittest.TestCase):
    """conversion_rules.xlsx phai dong bo voi JSON (dot 5 — documents/10 task 3)."""

    def test_xuat_xlsx_khop_json(self):
        import json
        from docs_exporter import export_rules_excel
        from openpyxl import load_workbook
        data = json.loads(_RULES_FILE.read_text(encoding="utf-8-sig"))
        out = Path(tempfile.mkdtemp()) / "rules_test.xlsx"
        export_rules_excel(str(_RULES_FILE), str(out))
        wb = load_workbook(str(out))
        self.assertEqual(wb.sheetnames, ["GioiThieu", "RuleVB", "RuleCS", "PairRules"])
        self.assertEqual(wb["RuleVB"].max_row - 3, len(data["vb_patterns"]))
        self.assertEqual(wb["RuleCS"].max_row - 3, len(data["cs_patterns"]))
        self.assertEqual(wb["PairRules"].max_row - 3, len(data["pair_rules"]))
        # id trong xlsx khop id trong json
        vb_ids_xlsx = {wb["RuleVB"].cell(r, 1).value for r in range(4, wb["RuleVB"].max_row + 1)}
        self.assertEqual(vb_ids_xlsx, {x["id"] for x in data["vb_patterns"]})

    def test_ban_xlsx_trong_repo_ton_tai(self):
        self.assertTrue((_RULES_FILE.parent / "conversion_rules.xlsx").is_file(),
                        "thieu rules/conversion_rules.xlsx — chay: python tool/docs_exporter.py")


class TestVbPatterns(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.rules = _rules()

    def _ids(self, body):
        return {n.split(":")[0].replace("RULE ", "") for n in check_vb_side(self.rules, body)}

    def test_cint_intdiv_iif(self):
        ids = self._ids("Dim x As Integer = CInt(a)\nDim y As Integer = a \\ b\nDim z = IIf(a > b, 1, 2)")
        self.assertEqual(ids, {"VB-CINT", "VB-INTDIV", "VB-IIF"})

    def test_onerror_rows0_emptystr(self):
        ids = self._ids('On Error Resume Next\nDim r = dt.Rows(0)\nIf s = "" Then\nEnd If')
        self.assertEqual(ids, {"VB-ONERROR", "VB-ROWS0", "VB-EMPTYSTR"})

    def test_andor_tran_nhung_khong_bat_andalso(self):
        self.assertIn("VB-ANDOR", self._ids("If a > 1 And b < 2 Then\nEnd If"))
        self.assertEqual(self._ids("If a > 1 AndAlso b < 2 OrElse c Then\nEnd If"), set())

    def test_khong_bat_trong_string_va_comment(self):
        # "AND" trong chuoi SQL va CInt trong comment khong duoc tinh
        ids = self._ids('Dim sql As String = "SELECT * FROM T WHERE A = 1 AND B = 2"\n\' dung CInt o day')
        self.assertEqual(ids, set())


class TestCsPatterns(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.rules = _rules()

    def _ids(self, body):
        return {n.split(":")[0].replace("RULE ", "") for n in check_cs_side(self.rules, body)}

    def test_wait_result_fireforget(self):
        self.assertEqual(self._ids("var x = task.Result; other.Wait();"),
                         {"CS-WAITRESULT"})
        self.assertEqual(self._ids("_ = CheckDuplicateAsync(code);"), {"CS-FIREFORGET"})

    def test_await_binh_thuong_khong_bat(self):
        self.assertEqual(self._ids("var x = await GetListAsync();"), set())


class TestPairRules(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.rules = _rules()

    def test_jp_msg_thieu_message(self):
        vb = _mk("RegisterX", 'MessageBox.Show("会員ランクが不正です。")\nMessageBox.Show("割引率が上限を超えています。")')
        cs = _mk("RegisterXAsync", 'return Result.Failure("割引率が上限を超えています。");')
        notes, escalate = check_pair_rules(self.rules, vb, cs)
        self.assertTrue(escalate)
        self.assertEqual(len([n for n in notes if "JP-MSG" in n]), 1)
        self.assertIn("会員ランクが不正です。", notes[0])

    def test_jp_msg_du_message_khong_bao(self):
        vb = _mk("RegisterX", 'MessageBox.Show("割引率が上限を超えています。")')
        cs = _mk("RegisterXAsync", 'return Result.Failure("割引率が上限を超えています。");')
        notes, escalate = check_pair_rules(self.rules, vb, cs)
        self.assertFalse(escalate)
        self.assertEqual(notes, [])

    def test_self_exclusion(self):
        vb = _mk("CheckDuplicateName", 'Dim sql As String = "SELECT 1 FROM T WHERE N = :n AND ID <> :id"')
        cs_thieu = _mk("CheckDuplicateNameAsync", "return await _context.T.AnyAsync(x => x.N == n);")
        cs_du = _mk("CheckDuplicateNameAsync", "return await _context.T.AnyAsync(x => x.N == n && x.Id != id);")
        notes, _ = check_pair_rules(self.rules, vb, cs_thieu)
        self.assertTrue(any("SELF-EXCL" in n for n in notes))
        notes, _ = check_pair_rules(self.rules, vb, cs_du)
        self.assertFalse(any("SELF-EXCL" in n for n in notes))
        # ten method khong phai check/exist/dup -> khong ap dung
        vb2 = _mk("GetName", vb.body)
        notes, _ = check_pair_rules(self.rules, vb2, cs_thieu)
        self.assertFalse(any("SELF-EXCL" in n for n in notes))

    def test_rownum_order(self):
        vb = _mk("GetTopOrder", 'Dim sql As String = "SELECT * FROM T WHERE ROWNUM <= 1"')
        cs_thieu = _mk("GetTopOrderAsync", "return await _context.T.FirstOrDefaultAsync();")
        cs_du = _mk("GetTopOrderAsync", "return await _context.T.OrderBy(x => x.Id).FirstOrDefaultAsync();")
        notes, _ = check_pair_rules(self.rules, vb, cs_thieu)
        self.assertTrue(any("ROWNUM-ORDER" in n for n in notes))
        notes, _ = check_pair_rules(self.rules, vb, cs_du)
        self.assertFalse(any("ROWNUM-ORDER" in n for n in notes))

    def test_validator_split(self):
        crit_lech = [CriterionResult("C5", "WARN", 12.0, 20.0)]
        crit_ok = [CriterionResult("C5", "OK", 20.0, 20.0)]
        with tempfile.TemporaryDirectory() as tmp:
            feature = Path(tmp) / "Features" / "X"
            feature.mkdir(parents=True)
            (feature / "XValidators.cs").write_text("// v", encoding="utf-8")
            note = validator_split_note(self.rules, crit_lech, r"Features\X\XCommands.cs", tmp)
            self.assertIn("VALIDATOR-SPLIT", note)
            # C4/C5 deu OK -> khong note; khong co file Validators -> khong note
            self.assertEqual(validator_split_note(self.rules, crit_ok, r"Features\X\XCommands.cs", tmp), "")
            self.assertEqual(validator_split_note(self.rules, crit_lech, r"Features\Y\YCommands.cs", tmp), "")


class TestIntegrationSamples(unittest.TestCase):
    """Rule chay tren bo sample PCRS — ket qua da chot trong documents/06."""

    @classmethod
    def setUpClass(cls):
        set_config(Config())
        root = Path(__file__).parent.parent
        cls.result = scan_folders(str(root / "samples" / "pcrs" / "legacy_vb"),
                                  str(root / "samples" / "pcrs" / "new_aspcore"))
        cls.by_name = {c.name: c for c in cls.result.comparisons}

    def _rule_ids(self, name):
        return {n.split(":")[0].replace("RULE ", "")
                for n in self.by_name[name].notes if n.startswith("RULE ")}

    def test_calc_rounded_price(self):
        comp = self.by_name["CalcRoundedPrice"]
        self.assertEqual(comp.status, "PASS")
        self.assertEqual(self._rule_ids("CalcRoundedPrice"), {"VB-CINT", "VB-INTDIV"})

    def test_check_duplicate_product_name(self):
        comp = self.by_name["CheckDuplicateProductName"]
        self.assertEqual(comp.status, "PASS")
        self.assertIn("SELF-EXCL", self._rule_ids("CheckDuplicateProductName"))

    def test_apply_member_discount_jp_msg(self):
        comp = self.by_name["ApplyMemberDiscount"]
        self.assertEqual(comp.status, "WARNING")
        self.assertIn("JP-MSG", self._rule_ids("ApplyMemberDiscount"))
        self.assertTrue(any("会員ランクが不正です。" in n for n in comp.notes))

    def test_update_stock_jp_msg(self):
        # 2 message tieng Nhat cua UpdateStock khong duoc convert (body viet lai)
        self.assertIn("JP-MSG", self._rule_ids("UpdateStock"))

    def test_dem_rule_hits(self):
        self.assertEqual(self.result.count_rule_hits(), 14)

    def test_tat_rules_qua_config(self):
        root = Path(__file__).parent.parent
        result = scan_folders(str(root / "samples" / "pcrs" / "legacy_vb"),
                              str(root / "samples" / "pcrs" / "new_aspcore"),
                              config=Config({"rules": {"enabled": False}}))
        set_config(Config())  # tra lai mac dinh cho test khac
        self.assertEqual(result.count_rule_hits(), 0)
        # tat rule khong duoc lam doi status
        by = {c.name: c for c in result.comparisons}
        self.assertEqual(by["CalcRoundedPrice"].status, "PASS")
        self.assertEqual(by["ApplyMemberDiscount"].status, "WARNING")


if __name__ == "__main__":
    unittest.main()
