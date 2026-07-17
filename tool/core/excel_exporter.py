"""Xuat ket qua danh gia migration ra file Excel (.xlsx) bang openpyxl.

Gom:
  - Summary: thong tin scan + thong ke theo trang thai
  - Detail : moi dong mot method, to mau theo trang thai
  - M001, M002, ... (dot 9): moi dong Detail mot sheet mo ta rieng — source VB
    va source ASP Core nam file nao, chu ky + toan bo body de nguoi review
    nhin bang mat; hyperlink 2 chieu voi sheet Detail.
"""

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

from .config import get_config
from .models import (STATUS_PASS, STATUS_WARNING, STATUS_FAIL,
                     STATUS_MISSING, STATUS_EXTRA,
                     AI_PASS, AI_WARNING, AI_NOT_RUN)

_STATUS_STYLE = {
    STATUS_PASS:    ("C6EFCE", "006100"),
    STATUS_WARNING: ("FFEB9C", "9C6500"),
    STATUS_FAIL:    ("FFC7CE", "9C0006"),
    STATUS_MISSING: ("D9D9D9", "404040"),
    STATUS_EXTRA:   ("DDEBF7", "1F4E78"),
}

# Dot 6: mau cot "Status AI danh gia" (PASS xanh / WARNING vang / chua chay xam)
_AI_STYLE = {
    AI_PASS:    _STATUS_STYLE[STATUS_PASS],
    AI_WARNING: _STATUS_STYLE[STATUS_WARNING],
    AI_NOT_RUN: ("D9D9D9", "404040"),
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


_LINK_FONT = Font(color="0563C1", underline="single")
_CODE_FONT = Font(name="Consolas", size=9)
_CODE_FILL = PatternFill("solid", fgColor="F5F5F5")
_SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")

# Dot 12: style muc "AI DE XUAT GIAI PHAP" trong sheet Mxxx — code de xuat
# highlight nen vang, rieng dong duoc sua (co "// FIX") dam mau + chu do
_SUGGEST_SECTION_FILL = PatternFill("solid", fgColor="FFD966")
_SUGGEST_CODE_FILL = PatternFill("solid", fgColor="FFF2CC")
_SUGGEST_FIX_FILL = PatternFill("solid", fgColor="FFE08A")
_SUGGEST_FIX_FONT = Font(name="Consolas", size=9, bold=True, color="9C0006")
_SUGGEST_OVERVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")  # o cot Detail


def _status_fill(status):
    bg, fg = _STATUS_STYLE.get(status, ("FFFFFF", "000000"))
    return PatternFill("solid", fgColor=bg), Font(color=fg, bold=True)


def method_sheet_name(row_no: int) -> str:
    """Ten sheet mo ta cua dong Detail thu row_no (M001, M002...) — ngan,
    duy nhat, khong dinh gioi han 31 ky tu / ky tu cam cua Excel."""
    return f"M{row_no:03d}"


def _related_key(name: str) -> str:
    """Khoa gom nhom method trung ten (bo hau to Async — quy uoc C#)."""
    key = name.lower()
    if key.endswith("async") and len(key) > 5:
        key = key[:-5]
    return key


def _build_related_map(comparisons) -> dict:
    """key -> [so thu tu dong Detail] de liet ke '1 VB -> nhieu C#'.

    Gom theo ca ten VB lan ten C# cua tung dong — method doi ten qua mapping
    thuoc ca 2 nhom nen cac ban trung ten o ca 2 phia deu duoc noi voi nhau.
    """
    related = {}
    for no, comp in enumerate(comparisons, start=1):
        for k in _comp_keys(comp):
            related.setdefault(k, []).append(no)
    return related


def _comp_keys(comp) -> set:
    """Tap khoa ten cua 1 dong: ten 2 phia + khoa khai bao tach 1-n (dot 11)."""
    keys = set()
    if comp.vb:
        keys.add(_related_key(comp.vb.name))
    if comp.cs:
        keys.add(_related_key(comp.cs.name))
    if not keys:
        keys.add(_related_key(comp.name))
    keys.update(_related_key(n) for n in comp.related_names)
    return keys


def _related_rows(comp, no, related_map) -> list:
    """Cac dong Detail khac trung khoa ten voi dong `no` (da bo chinh no)."""
    rows = set()
    for k in _comp_keys(comp):
        rows.update(related_map.get(k, []))
    rows.discard(no)
    return sorted(rows)


def _write_header(ws, row, headers, widths=None):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    if widths:
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w


# ---------- Sheet mo ta chi tiet tung method (dot 9) ----------

def _ws_link(ws, row, col, text, location):
    """Ghi 1 o hyperlink noi bo (nhay giua cac sheet trong file)."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=location)
    cell.font = _LINK_FONT
    return cell


def _write_info_row(ws, row, label, value):
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value=value)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return row + 1


def _write_section_header(ws, row, text, fill=_SECTION_FILL):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=11)
    cell.fill = fill
    ws.cell(row=row, column=2).fill = fill
    return row + 1


def _write_code_block(ws, row, body: str, max_lines: int) -> int:
    """In body code — moi dong code 1 dong Excel, font Consolas, nen xam nhat."""
    lines = (body or "").splitlines() or ["(body rong)"]
    if body and not body.strip():
        lines = ["(body rong)"]
    shown = lines[:max_lines]
    for line in shown:
        cell = ws.cell(row=row, column=2, value=line.rstrip()[:1000])
        cell.font = _CODE_FONT
        cell.fill = _CODE_FILL
        row += 1
    if len(lines) > max_lines:
        cell = ws.cell(row=row, column=2,
                       value=f"... (con {len(lines) - max_lines} dong — xem file goc)")
        cell.font = Font(italic=True, color="808080")
        row += 1
    return row


def _write_suggestion_code_block(ws, row, code: str, max_lines: int) -> int:
    """In code AI de xuat (dot 12) — nen vang de phan biet voi source goc;
    dong duoc AI sua (ket thuc bang '// FIX: ...') to dam + chu do."""
    lines = (code or "").splitlines()
    shown = lines[:max_lines]
    for line in shown:
        cell = ws.cell(row=row, column=2, value=line.rstrip()[:1000])
        if "// FIX" in line:
            cell.font = _SUGGEST_FIX_FONT
            cell.fill = _SUGGEST_FIX_FILL
        else:
            cell.font = _CODE_FONT
            cell.fill = _SUGGEST_CODE_FILL
        row += 1
    if len(lines) > max_lines:
        cell = ws.cell(row=row, column=2,
                       value=f"... (con {len(lines) - max_lines} dong)")
        cell.font = Font(italic=True, color="808080")
        row += 1
    return row


def _write_suggestion_section(ws, row, comp, max_lines) -> int:
    """Muc 'AI DE XUAT GIAI PHAP' trong sheet Mxxx (dot 12) — chi xuat hien khi
    AI cham WARNING va co de xuat: tom tat + giai thich chi tiet + code da sua."""
    row = _write_section_header(
        ws, row, "AI DE XUAT GIAI PHAP (cho dong AI cham WARNING)",
        fill=_SUGGEST_SECTION_FILL)
    if comp.ai_suggestion:
        row = _write_info_row(ws, row, "Tom tat huong sua", comp.ai_suggestion)
    if comp.ai_suggestion_detail:
        row = _write_info_row(ws, row, "Giai thich chi tiet", comp.ai_suggestion_detail)
    if comp.ai_suggestion_code:
        ws.cell(row=row, column=1, value="Code de xuat").font = Font(bold=True)
        row = _write_suggestion_code_block(ws, row, comp.ai_suggestion_code, max_lines)
        legend = ws.cell(row=row, column=2,
                         value="(Dong nen dam co '// FIX:' la dong AI sua/them — "
                               "doi chieu voi SOURCE ASP.NET CORE o duoi truoc khi ap dung)")
        legend.font = Font(italic=True, color="808080")
        row += 1
    else:
        cell = ws.cell(row=row, column=2,
                       value="(AI khong de xuat code — xem giai thich chi tiet o tren, "
                             "can nguoi review xac nhan tay)")
        cell.font = Font(italic=True, color="808080")
        row += 1
    return row + 1


def _write_method_source(ws, row, side_label, info, absent_text, max_lines) -> int:
    """Mot muc SOURCE (VB hoac ASP Core): file/dong/chu ky/body — hoac ly do vang mat."""
    row = _write_section_header(ws, row, side_label)
    if info is None:
        cell = ws.cell(row=row, column=2, value=absent_text)
        cell.font = Font(bold=True, color="9C0006")
        return row + 2
    row = _write_info_row(ws, row, "File", f"{info.file} : dong {info.line}")
    row = _write_info_row(ws, row, "Chu ky", info.signature)
    ws.cell(row=row, column=1, value="Body").font = Font(bold=True)
    row = _write_code_block(ws, row, info.body, max_lines)
    return row + 1


def _write_method_sheet(wb, no, comp, related_map, comparisons, max_lines):
    """Sheet Mxxx — mo ta day du 1 dong Detail de nguoi review danh gia bang mat."""
    ws = wb.create_sheet(method_sheet_name(no))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 120
    ws.sheet_view.showGridLines = False

    # Tieu de + link quay lai dung dong Detail
    fill, font = _status_fill(comp.status)
    title = ws.cell(row=1, column=1,
                    value=f"{method_sheet_name(no)} — {comp.name}   [{comp.status}]")
    title.font = Font(bold=True, size=13)
    ws.cell(row=1, column=2).fill = fill
    ws.cell(row=1, column=2, value=comp.status).font = font
    ws.cell(row=1, column=2).fill = fill
    _ws_link(ws, 2, 1, "<- Quay lai sheet Detail", f"Detail!B{no + 1}")

    # Tom tat
    row = 4
    if comp.vb and comp.cs:
        row = _write_info_row(ws, row, "Score / Similarity",
                              f"{comp.score:.0f}/100   ·   {comp.similarity:.2f}")
    if comp.matched_by_mapping:
        row = _write_info_row(ws, row, "Mapping",
                              f"Khop qua bang mapping ten thu cong ('{comp.vb.name}' -> '{comp.cs.name}')")
    if comp.ai_status:
        row = _write_info_row(ws, row, "AI danh gia",
                              f"[{comp.ai_status}] {comp.ai_comment}")
    if comp.notes:
        row = _write_info_row(ws, row, "Notes", "\n".join(f"- {n}" for n in comp.notes))
    row += 1

    # Dot 12: AI de xuat giai phap — chi khi AI cham WARNING va co noi dung
    if comp.ai_suggestion or comp.ai_suggestion_detail or comp.ai_suggestion_code:
        row = _write_suggestion_section(ws, row, comp, max_lines)

    # 2 muc source — cover du: cap 1-1, MISSING (chua implement), EXTRA (viet moi)
    row = _write_method_source(
        ws, row, "SOURCE VB (HE THONG CU)", comp.vb,
        "KHONG CO o source VB — method viet moi tren he thong ASP.NET Core (EXTRA)",
        max_lines)
    row = _write_method_source(
        ws, row, "SOURCE ASP.NET CORE (HE THONG MOI)", comp.cs,
        "CHUA IMPLEMENT tren ASP.NET Core — method chi ton tai o source VB (MISSING); "
        "xem Notes o tren (UI event co the da chuyen hop le sang .razor)",
        max_lines)

    # Method lien quan trung ten — the hien '1 VB -> nhieu C# o nhieu folder'
    others = _related_rows(comp, no, related_map)
    if others:
        row = _write_section_header(
            ws, row, "METHOD LIEN QUAN TRUNG TEN (1 method VB co the ung voi nhieu method C#)")
        for other_no in others:
            other = comparisons[other_no - 1]
            side = "VB + C#" if (other.vb and other.cs) else ("chi VB" if other.vb else "chi C#")
            files = " / ".join(x.file for x in (other.vb, other.cs) if x)
            _ws_link(ws, row, 1, method_sheet_name(other_no),
                     f"{method_sheet_name(other_no)}!A1")
            ws.cell(row=row, column=2,
                    value=f"{other.name}   [{other.status}]   ({side})   {files}")
            row += 1


def export_excel(result, out_path: str):
    """Ghi ScanResult ra file .xlsx."""
    cfg = get_config()
    method_sheets_on = bool(cfg.excel("method_sheets")) and bool(result.comparisons)
    max_body_lines = int(cfg.excel("max_body_lines"))
    wb = Workbook()

    # ---------- Sheet Summary ----------
    ws = wb.active
    ws.title = "Summary"
    counts = result.count_by_status()
    vb_total = sum(counts[s] for s in (STATUS_PASS, STATUS_WARNING, STATUS_FAIL, STATUS_MISSING))
    pass_rate = (counts[STATUS_PASS] / vb_total * 100) if vb_total else 0

    ws["A1"] = "BAO CAO KIEM TRA MIGRATION BACKEND (VB.NET -> ASP.NET Core)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    info_rows = [
        ("Thoi gian scan", result.scanned_at),
        ("Folder VB (he thong cu)", result.vb_folder),
        ("Folder C# (he thong moi)", result.cs_folder),
        ("Folder Business Logic (C#) da quet",
         f"{result.cs_scan_folder or result.cs_folder} — {result.cs_detect_note}"
         if result.cs_detect_note else (result.cs_scan_folder or result.cs_folder)),
        ("So file VB / C#", f"{result.vb_file_count} / {result.cs_file_count}"),
        ("Tong so method phia VB", vb_total),
        ("Diem trung binh (cac method map duoc)", result.average_score()),
        ("Ty le PASS", f"{pass_rate:.1f}%"),
    ]
    if result.mapping_file:
        info_rows.append(("File mapping ten method (thu cong)", result.mapping_file))
    if result.count_frontend_handled():
        info_rows.append(("MISSING la UI event da xac nhan co handler frontend (.razor)",
                          result.count_frontend_handled()))
    if result.count_rule_hits():
        info_rows.append(("Method co ghi chu RULE can review (rules/conversion_rules.json)",
                          result.count_rule_hits()))
    if result.ai_reviewed():
        ai = result.count_ai()
        info_rows.append(("AI danh gia (Claude API — lop danh gia doc lap, khong anh huong diem C1-C5)",
                          f"PASS: {ai[AI_PASS]}  WARNING: {ai[AI_WARNING]}  "
                          f"Chua danh gia: {ai['not_run']}"))
        info_rows.append(("Cot 'AI de xuat giai phap' (sheet Detail — dot 12)",
                          "Tom tat huong sua cho dong AI cham WARNING; giai thich chi tiet "
                          "+ code de xuat (highlight vang, dong sua danh dau '// FIX:') "
                          "nam trong sheet mo ta Mxxx cua tung dong"))
    info_rows.append(("Cot 'Status DEV danh gia' (sheet Detail)",
                      "Nguoi review tu cham bang dropdown (PASS/WARNING/FAIL/MISSING/EXTRA) "
                      "— tool va AI khong ghi vao cot nay"))
    if method_sheets_on:
        info_rows.append(("Sheet mo ta chi tiet tung method (M001...)",
                          f"Click ten method o cot B sheet Detail de nhay toi sheet mo ta "
                          f"(source VB + ASP Core: file, chu ky, body day du) — "
                          f"{len(result.comparisons)} sheet"))
    row = 3
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    _write_header(ws, row, ["Trang thai", "So luong", "Y nghia"], [22, 12, 70])
    extra_meaning = "Chi co o C# — method viet moi (ghi nhan, khong phai loi)"
    new_arch = result.count_new_arch()
    if new_arch:
        extra_meaning += (f"; trong do {new_arch} thuoc kien truc moi "
                          "(MediatR Handler / DTO / Result / Validator — documents/04, xem cot Notes sheet Detail)")
    meanings = {
        STATUS_PASS: "Convert dung (diem >= 85, khong tieu chi nao fail)",
        STATUS_WARNING: "Can review tay (co canh bao hoac diem 60-85)",
        STATUS_FAIL: "Nhieu kha nang convert sai (sai chu ky / logic lech lon / diem < 60)",
        STATUS_MISSING: "Chi co o VB — chua convert, bi loai bo, hoac da chuyen sang frontend",
        STATUS_EXTRA: extra_meaning,
    }
    for status in (STATUS_PASS, STATUS_WARNING, STATUS_FAIL, STATUS_MISSING, STATUS_EXTRA):
        row += 1
        fill, font = _status_fill(status)
        c = ws.cell(row=row, column=1, value=status)
        c.fill, c.font, c.border = fill, font, _BORDER
        ws.cell(row=row, column=2, value=counts[status]).border = _BORDER
        ws.cell(row=row, column=3, value=meanings[status]).border = _BORDER
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 55

    # ---------- Sheet Detail ----------
    ws = wb.create_sheet("Detail")
    headers = ["No", "Method", "VB File", "C# File", "VB Signature", "C# Signature",
               "C1 Ton tai", "C2 Tham so", "C3 Kieu tra ve", "C4 Cau truc",
               "C5 Logic", "Similarity", "Score", "Status", "Notes",
               "Nội dung AI đánh giá", "Status AI đánh giá",
               "AI đề xuất giải pháp", "Status DEV đánh giá"]
    widths = [5, 26, 24, 30, 45, 45, 10, 10, 12, 10, 10, 10, 8, 11, 60, 60, 24, 55, 20]
    _write_header(ws, 1, headers, widths)
    ws.freeze_panes = "C2"

    def crit_label(comp, code):
        for c in comp.criteria:
            if c.code == code:
                return c.label
        return "-"

    for i, comp in enumerate(result.comparisons, start=1):
        r = i + 1
        fill, font = _status_fill(comp.status)
        # Dot 6: 2 cot AI — chua chay/loi -> "AI chua thuc hien danh gia"
        ai_status = comp.ai_status if comp.ai_status in (AI_PASS, AI_WARNING) else AI_NOT_RUN
        # Dot 12: cot "AI de xuat giai phap" — CHI tom tat (overview); giai thich
        # chi tiet + code de xuat nam trong sheet mo ta Mxxx cua dong nay
        ai_sugg = comp.ai_suggestion
        if ai_sugg and method_sheets_on and (comp.ai_suggestion_detail
                                             or comp.ai_suggestion_code):
            ai_sugg += (f"\n(Giai thich chi tiet + code de xuat: "
                        f"sheet {method_sheet_name(i)})")
        values = [
            i, comp.name,
            comp.vb.file if comp.vb else "-",
            comp.cs.file if comp.cs else "-",
            comp.vb.signature if comp.vb else "-",
            comp.cs.signature if comp.cs else "-",
            crit_label(comp, "C1"), crit_label(comp, "C2"), crit_label(comp, "C3"),
            crit_label(comp, "C4"), crit_label(comp, "C5"),
            comp.similarity if comp.vb and comp.cs else "-",
            comp.score if comp.vb and comp.cs else "-",
            comp.status, comp.note_text,
            comp.ai_comment, ai_status, ai_sugg,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(col in (5, 6, 15, 16, 18)))
        if method_sheets_on:
            # Dot 9: o ten method la hyperlink nhay sang sheet mo ta Mxxx
            name_cell = ws.cell(row=r, column=2)
            name_cell.hyperlink = Hyperlink(
                ref=name_cell.coordinate, location=f"{method_sheet_name(i)}!A1")
            name_cell.font = _LINK_FONT
        status_cell = ws.cell(row=r, column=14)
        status_cell.fill, status_cell.font = fill, font
        ai_bg, ai_fg = _AI_STYLE.get(ai_status, ("FFFFFF", "000000"))
        ai_cell = ws.cell(row=r, column=17)
        ai_cell.fill = PatternFill("solid", fgColor=ai_bg)
        ai_cell.font = Font(color=ai_fg, bold=(ai_status != AI_NOT_RUN))
        if ai_sugg:
            # o co de xuat -> nen vang nhat de dap vao mat nguoi review
            ws.cell(row=r, column=18).fill = _SUGGEST_OVERVIEW_FILL
        ws.cell(row=r, column=19).border = _BORDER  # o DEV de trong cho reviewer chon

    last_row = len(result.comparisons) + 1
    ws.auto_filter.ref = f"A1:S{last_row}"

    # Dot 7 (dot 12 doi cot R -> S): cot "Status DEV danh gia" — dropdown cho
    # nguoi review tu cham, gia tri cung bo trang thai voi cot Status cua tool
    if result.comparisons:
        dev_statuses = [STATUS_PASS, STATUS_WARNING, STATUS_FAIL,
                        STATUS_MISSING, STATUS_EXTRA]
        dv = DataValidation(
            type="list", formula1=f'"{",".join(dev_statuses)}"', allow_blank=True,
            errorTitle="Gia tri khong hop le",
            error="Chon mot trong: " + ", ".join(dev_statuses))
        ws.add_data_validation(dv)
        dev_range = f"S2:S{last_row}"
        dv.add(dev_range)
        # To mau o theo gia tri reviewer chon (conditional formatting)
        for status in dev_statuses:
            bg, fg = _STATUS_STYLE[status]
            ws.conditional_formatting.add(dev_range, CellIsRule(
                operator="equal", formula=[f'"{status}"'],
                fill=PatternFill("solid", fgColor=bg),
                font=Font(color=fg, bold=True)))

    # ---------- Sheet mo ta tung method (dot 9) ----------
    if method_sheets_on:
        related_map = _build_related_map(result.comparisons)
        for no, comp in enumerate(result.comparisons, start=1):
            _write_method_sheet(wb, no, comp, related_map,
                                result.comparisons, max_body_lines)

    wb.save(out_path)
    return out_path
