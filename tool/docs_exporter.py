# -*- coding: utf-8 -*-
"""Xuat tai lieu dang Excel (dot 5 — documents/10):

  1. rules/conversion_rules.xlsx        — ban Excel cua file rule (nguon: JSON)
  2. documents/11_Tool_Usage_Guideline.xlsx — guideline su dung tool

JSON van la NGUON SU THAT (tool doc JSON khi chay); XLSX chi de trinh bay cho
nguoi review. Sua rules/conversion_rules.json xong chay lai:

    python tool/docs_exporter.py
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).parent.parent

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11, color="1F4E78")
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="top", wrap_text=True)


def _write_header(ws, row, headers, widths=None):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    if widths:
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _write_rows(ws, start_row, rows):
    r = start_row
    for values in rows:
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = _WRAP
        r += 1
    return r


def _title(ws, text, span):
    ws["A1"] = text
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)


# ====================================================================
# 1. rules/conversion_rules.json -> rules/conversion_rules.xlsx
# ====================================================================

def export_rules_excel(json_path: str, xlsx_path: str) -> str:
    data = json.loads(Path(json_path).read_text(encoding="utf-8-sig"))
    wb = Workbook()

    # ---- Sheet GioiThieu ----
    ws = wb.active
    ws.title = "GioiThieu"
    _title(ws, "TIEU CHI DANH GIA CHAT LUONG CONVERT — RULE CHECK", 2)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    rows = [
        ("Phien ban / cap nhat", f"{data.get('version', '')} / {data.get('updated', '')}"),
        ("Gioi thieu", data.get("_gioi_thieu", "")),
        ("Cach doc", data.get("_cach_doc", "")),
        ("Luu y dong bo", "File nay duoc SINH TU rules/conversion_rules.json — JSON la nguon su that "
                          "(tool doc JSON khi chay). Sua JSON xong chay: python tool/docs_exporter.py "
                          "de tai sinh file Excel nay."),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = _WRAP
        ws.cell(row=r, column=2, value=value).alignment = _WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Tieu chi cham diem C1–C5 + SQL (documents/03)").font = _SECTION_FONT
    r += 1
    _write_header(ws, r, ["Tieu chi", "Noi dung"])
    r = _write_rows(ws, r + 1, sorted(data.get("_tieu_chi_C1_C5", {}).items()))

    # ---- Sheet RuleVB / RuleCS ----
    for sheet, key, label in (("RuleVB", "vb_patterns", "body VB (he thong cu)"),
                              ("RuleCS", "cs_patterns", "body C# (he thong moi)")):
        ws = wb.create_sheet(sheet)
        _title(ws, f"RULE QUET {label.upper()}", 5)
        _write_header(ws, 3, ["ID", "Tieu de", "Regex", "Vi sao nguy hiem", "Can kiem tra gi"],
                      [16, 34, 28, 60, 60])
        _write_rows(ws, 4, [(x["id"], x["title"], x["pattern"], x["detail"], x["advice"])
                            for x in data.get(key, [])])
        ws.freeze_panes = "A4"

    # ---- Sheet PairRules ----
    ws = wb.create_sheet("PairRules")
    _title(ws, "RULE SO SANH 2 CHIEU TREN MOT CAP METHOD", 5)
    _write_header(ws, 3, ["ID", "Ten rule", "Dang bat", "Mo ta", "Hanh dong khi trung"],
                  [18, 22, 10, 80, 34])
    pair_rows = []
    for name, cfg in data.get("pair_rules", {}).items():
        action = "Them note RULE (chi review, khong doi status)"
        if cfg.get("escalate_pass_to_warning"):
            action = "Them note + NANG PASS -> WARNING (nghi van loi that)"
        pair_rows.append((cfg.get("id", name), cfg.get("title", ""),
                          "Bat" if cfg.get("enabled") else "Tat",
                          cfg.get("detail", ""), action))
    _write_rows(ws, 4, pair_rows)
    ws.freeze_panes = "A4"

    wb.save(xlsx_path)
    return xlsx_path


# ====================================================================
# 2. documents/11_Tool_Usage_Guideline.xlsx
# ====================================================================

def export_guideline_excel(xlsx_path: str) -> str:
    wb = Workbook()

    # ---- TongQuan ----
    ws = wb.active
    ws.title = "TongQuan"
    _title(ws, "MIGRATION CHECKER — GUIDELINE SU DUNG (VB.NET -> ASP.NET Core)", 2)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110
    rows = [
        ("Tool lam gi", "Kiem tra chat luong migration MUC METHOD giua he thong cu (VB.NET/.NET Framework 4.8/"
                        "Oracle 19c) va he thong moi (C#/ASP.NET Core/PostgreSQL 17, kien truc CQRS 6 project theo "
                        "documents/04). Cham diem C1-C5 tung cap method, kiem tra SQL Oracle sot, xac nhan UI event "
                        "sang .razor, quet rule cac bay convert, gate danh gia AI qua Claude API (dot 6 — tuy chon) "
                        "— xuat bao cao Excel (kem cot 'Status DEV danh gia' cho nguoi review tu cham, dot 7)."),
        ("Cach chay 1 — exe (khuyen dung)", "Double-click dist\\MigrationChecker.exe (khong can cai Python). "
                                            "File config.json va thu muc rules\\ dat canh exe se duoc tu doc."),
        ("Cach chay 2 — tu source", "pip install openpyxl  ->  python tool\\main.py (can Python 3.10+)."),
        ("Cach chay 3 — CLI (CI/CD, retest)", "python tool\\cli.py <folder_VB> <folder_ASPCore> [tuy chon] — "
                                              "xem sheet CLI_CICD."),
        ("Bo sample kem theo", "samples\\pcrs\\legacy_vb (30 method VB) + samples\\pcrs\\new_aspcore (solution PCRS) "
                               "+ samples\\pcrs\\method_mapping.csv. Ky vong ket qua: documents/06."),
        ("Tieu chi rule check", "rules\\conversion_rules.json (tool doc) + rules\\conversion_rules.xlsx (ban doc cho "
                                "nguoi review). Sua JSON khong can build lai tool."),
        ("Bo test tu dong", "python -m unittest discover tests — PHAI chay sau moi lan sua engine."),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = _WRAP
        ws.cell(row=r, column=2, value=value).alignment = _WRAP
        r += 1

    # ---- QuyTrinh ----
    ws = wb.create_sheet("QuyTrinh")
    _title(ws, "QUY TRINH 5 BUOC TREN GUI", 3)
    _write_header(ws, 3, ["Buoc", "Thao tac", "Ghi chu"], [8, 52, 80])
    _write_rows(ws, 4, [
        (1, "Chon folder VB.NET (he thong cu)", "Chon FOLDER GOC chua *.vb — business logic VB nam trong "
                                                "code-behind nen tool luon quet toan bo."),
        (2, "Chon folder ASP.NET Core (he thong moi)", "Chon FOLDER GOC cua ca solution — tool tu tim project "
                                                       "*Application* chua Features/ (Business Logic) de quet."),
        (3, "(Tuy chon) Chon file mapping ten method", "Chi can khi co method bi DOI TEN khi migrate. Dinh dang CSV: "
                                                       "TenVB,TenCSharp — xem samples\\pcrs\\method_mapping.csv."),
        (4, "Bam Scan, doc ket qua", "Loc theo status (uu tien FAIL truoc, WARNING sau), go ten vao o 'Tim method'. "
                                     "Click 1 dong -> tab 'Tong quan' (tieu chi, ghi chu) va tab 'Code VB <-> C#' "
                                     "(body 2 ben canh nhau)."),
        (5, "Export Excel", "Xuat bao cao .xlsx gui team review (Summary + Detail + sheet mo ta M001... — "
                            "dot 9: click ten method o cot B sheet Detail de xem source VB va ASP Core "
                            "day du body ngay trong Excel). Sheet Detail co 3 cot status: Status (may cham "
                            "C1-C5) -> 'Status AI danh gia' (AI goi y — dot 6) -> 'Status DEV danh gia' "
                            "(dropdown de nguoi review chon PASS/WARNING/FAIL/MISSING/EXTRA lam ket luan "
                            "cuoi — dot 7). Dong AI cham WARNING co them cot 'AI de xuat giai phap' "
                            "(tom tat huong sua — dot 12); giai thich chi tiet + code de xuat highlight "
                            "vang nam trong sheet mo ta Mxxx cua dong do."),
    ])
    r = 10
    ws.cell(row=r, column=1, value="Thanh phan man hinh").font = _SECTION_FONT
    r += 1
    _write_header(ws, r, ["Thanh phan", "Y nghia", ""], [30, 52, 80])
    _write_rows(ws, r + 1, [
        ("Business Logic (tu phat hien)", "Thu muc C# tool thuc su quet, hien sau khi Scan", ""),
        ("Loc trang thai / Tim method", "Loc bang ket qua theo status va theo ten (ket hop duoc)", ""),
        ("Dong tong ket", "Dem PASS/WARNING/FAIL/MISSING/EXTRA + diem trung binh + so method co ghi chu RULE", ""),
        ("Panel chi tiet — tab Tong quan", "Chu ky 2 ben (file:dong), nhan C1-C5, similarity, toan bo ghi chu", ""),
        ("Panel chi tiet — tab Code VB <-> C#", "Body method 2 he thong dat canh nhau de doi chieu bang mat", ""),
    ])

    # ---- TrangThai ----
    ws = wb.create_sheet("TrangThai")
    _title(ws, "TRANG THAI TONG HOP + TIEU CHI CHAM DIEM", 3)
    _write_header(ws, 3, ["Trang thai", "Y nghia", "Huong xu ly"], [14, 64, 64])
    _write_rows(ws, 4, [
        ("PASS", "Convert dung (diem >= 85, khong tieu chi nao fail)", "Van doc ghi chu RULE neu co — PASS co the "
                 "mang diem ngam can xac nhan (vd CInt lam tron)."),
        ("WARNING", "Can review tay (co canh bao hoac diem 60-85)", "Doc Notes: phan lon la 'thay doi co chu dich' "
                    "cua kien truc moi, xac nhan nhanh la xong."),
        ("FAIL", "Nhieu kha nang convert sai (sai chu ky / logic lech lon / diem < 60)", "BAT BUOC doi chieu code "
                 "2 ben tai file:dong trong panel chi tiet."),
        ("MISSING", "Chi co o VB — chua convert / bi loai / da chuyen frontend", "Neu ghi chu 'da tim thay handler "
                    "... .razor' thi da chuyen hop le; con lai phai xac nhan ly do."),
        ("EXTRA", "Chi co o C# — method viet moi", "Ghi chu 'EXTRA (kien truc moi)' (Handler/DTO/Result/Validator) "
                  "thi bo qua — khong phai thua code."),
        ("Status AI danh gia (dot 6)", "PASS/WARNING do AI cham (Claude tra phi hoac Gemini free tier — tool tu "
         "nhan dien theo key, dot 8) — lop goi y doc lap, khong anh huong diem C1-C5; "
         "'AI chua thuc hien danh gia' = chua bat llm/thieu token/loi API",
         "Doc cot 'Noi dung AI danh gia' ben canh; dong WARNING doc them cot 'AI de xuat giai phap' "
         "(tom tat huong sua — dot 12) va sheet mo ta Mxxx (giai thich chi tiet + code de xuat, dong "
         "sua danh dau '// FIX:'). Bat AI: config.json muc llm (enabled + api_key — "
         "key AIza... lay free tai aistudio.google.com, key sk-ant... tai platform.claude.com)."),
        ("Status DEV danh gia (dot 7)", "Cot danh cho NGUOI REVIEW tu cham ket luan cuoi cung — o trong, "
         "co dropdown PASS/WARNING/FAIL/MISSING/EXTRA, tu to mau theo gia tri chon",
         "Sau khi doi chieu Status (may) va Status AI, chon gia tri cuoi cung vao day roi luu file."),
    ])
    r = 12
    ws.cell(row=r, column=1, value="Tieu chi C1-C5").font = _SECTION_FONT
    r += 1
    _write_header(ws, r, ["Ma", "Kiem tra", "Trong so"], [14, 64, 64])
    _write_rows(ws, r + 1, [
        ("C1", "Method ton tai o ca 2 ben (khop ten / hau to Async / bang mapping)", 30),
        ("C2", "So luong + kieu tham so (DataTable -> List<DTO> chi canh bao)", 20),
        ("C3", "Kieu tra ve (Task<X> ~ X; Boolean -> Result chi canh bao)", 15),
        ("C4", "Cau truc dieu khien (if / vong lap / try / return / throw)", 15),
        ("C5", "Do tuong dong logic sau chuan hoa token (>=0.75 OK, 0.5-0.75 WARN, <0.5 NG)", 20),
    ])

    # ---- GhiChu ----
    ws = wb.create_sheet("GhiChu")
    _title(ws, "CACH DOC CAC GHI CHU (NOTES)", 3)
    _write_header(ws, 3, ["Ghi chu bat dau bang", "Y nghia", "Xu ly"], [36, 62, 62])
    _write_rows(ws, 4, [
        ("C1: khop ten qua hau to Async / khop qua bang mapping", "Cach tool ghep cap method", "Khong phai loi."),
        ("C2/C3: DataTable -> List<...> / Result pattern", "Thay doi co chu dich cua kien truc moi (documents/04)",
         "Xac nhan nhanh, khong phai bug."),
        ("C5: soft delete / viet lai theo EF Core / khac thu tu cau lenh", "Khoan dung thay doi co chu dich",
         "Review nhanh de xac nhan."),
        ("SQL: con cu phap Oracle (...)", "SQL phia C# con NVL/ROWNUM/SYSDATE/DUAL/.NEXTVAL/DECODE/(+) — "
         "SE LOI tren PostgreSQL", "LOI THAT — phai sua code C#. Tool ep toi thieu WARNING."),
        ("UI event — da tim thay handler ... .razor", "UI event da co diem den ben frontend", "Da chuyen hop le."),
        ("RULE <id>: ...", "Diem can review tay theo rules/conversion_rules.xlsx (bay ngu nghia VB->C#, "
         "code smell async, v.v.) — khong tru diem", "Mo rules/conversion_rules.xlsx doc cot 'Can kiem tra gi' "
         "cua rule tuong ung."),
        ("RULE JP-MSG: message '...' khong thay ben C#", "Message tieng Nhat cua MessageBox/VB khong xuat hien o "
         "Result.Failure ben C# — kha nang THIEU nhanh check", "Nghi van loi that — PASS bi nang WARNING; doi chieu "
         "tung nhanh if cua method cu."),
        ("RULE VALIDATOR-SPLIT", "C4/C5 lech nhung cung thu muc co *Validators.cs — mot phan check co the da tach "
         "sang FluentValidation", "Doi chieu file Validators truoc khi ket luan thieu logic."),
    ])

    # ---- CLI_CICD ----
    ws = wb.create_sheet("CLI_CICD")
    _title(ws, "CLI — CHAY TU DONG / GAN VAO CI-CD", 2)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 100
    _write_header(ws, 3, ["Tham so", "Y nghia"])
    _write_rows(ws, 4, [
        ("<folder_VB> <folder_ASPCore>", "2 tham so bat buoc: folder goc 2 he thong."),
        ("--out report_{ts}.xlsx", "Xuat bao cao Excel; token {ts} tu thay bang timestamp YYYYMMDD_HHMMSS "
         "de cac lan xuat khong trung ten (ap dung ca --json). GUI: ten mac dinh da kem timestamp."),
        ("--map mapping.csv", "Bang mapping ten method bi doi ten (CSV 'TenVB,TenCSharp' hoac JSON). "
         "Dot 11 — method bi TACH thanh nhieu method ten khac: them cot "
         "'TenVB,TenCSharpChinh,ManhTach1,...' (manh tach nhan note 'EXTRA (manh tach)' "
         "va duoc gom vao muc METHOD LIEN QUAN cua sheet mo ta)."),
        ("--config config.json", "Ghi de trong so/nguong/khoan dung (xem config.sample.json)."),
        ("--rules file.json", "Thay file rule check (mac dinh tu tim rules/conversion_rules.json)."),
        ("--json out.json", "Xuat ket qua JSON lam baseline cho lan scan sau."),
        ("--baseline truoc.json", "So voi lan scan truoc: in bang tot len / xau di / moi / bien mat."),
        ("--fail-on FAIL,MISSING", "Exit code 2 neu con method thuoc status liet ke (EXTRA kien truc moi va "
         "MISSING da xac nhan frontend khong tinh)."),
        ("--min-score 85", "Exit code 2 neu diem trung binh duoi nguong."),
        ("--no-detect", "Tat auto-detect Business Logic, quet toan bo folder C#."),
        ("--llm", "Bat gate danh gia AI qua Claude API (dot 6) — them cac cot "
         "'Noi dung AI danh gia' / 'Status AI danh gia' / 'AI de xuat giai phap' (dot 12) vao Excel; "
         "dong AI cham WARNING co giai thich + code de xuat trong sheet mo ta Mxxx; "
         "can pip install anthropic + API key trong config.json (muc llm) hoac bien moi truong "
         "ANTHROPIC_API_KEY; khong anh huong diem C1-C5 va exit code."),
    ])
    r = 16
    ws.cell(row=r, column=1, value="Exit code: 0 = dat · 1 = loi input · 2 = khong dat nguong").font = Font(bold=True)
    r += 2
    ws.cell(row=r, column=1, value="Vi du lenh retest bo sample:").font = _SECTION_FONT
    ws.cell(row=r + 1, column=1,
            value="python tool/cli.py samples/pcrs/legacy_vb samples/pcrs/new_aspcore "
                  "--map samples/pcrs/method_mapping.csv --out samples/migration_report_pcrs.xlsx "
                  "--json samples/migration_report_pcrs.json").alignment = _WRAP
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=2)

    # ---- FAQ ----
    ws = wb.create_sheet("FAQ")
    _title(ws, "LOI THUONG GAP & FAQ", 2)
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 95
    _write_header(ws, 3, ["Tinh huong", "Nguyen nhan / cach xu ly"])
    _write_rows(ws, 4, [
        ("Bao 'Folder khong hop le'", "Duong dan khong ton tai hoac chua chon — chon lai bang nut Chon..."),
        ("Scan xong 0 method phia C#", "Folder khong chua .cs hoac Business Logic phat hien sai — xem dong "
         "'Business Logic (tu phat hien)' tren GUI."),
        ("Method da convert nhung bao MISSING + EXTRA cung nghia", "Method bi DOI TEN — khai bao vao file mapping "
         "(TenVB,TenCSharp) roi Scan lai."),
        ("Nhieu dong MISSING ten xxx_Click / xxx_Load", "UI event WinForms — neu co ghi chu 'da tim thay handler' "
         "la da chuyen hop le sang .razor."),
        ("Nhieu dong EXTRA ten Handle", "MediatR Handler cua CQRS — da danh dau 'kien truc moi', khong phai loi."),
        ("Muon doi trong so / nguong / khoan dung", "Copy config.sample.json thanh config.json dat canh exe, "
         "sua gia tri, Scan lai."),
        ("Muon them/sua rule check", "Sua rules/conversion_rules.json (Notepad) — khong can build lai; chay "
         "python tool/docs_exporter.py de tai sinh ban Excel."),
        ("Ket qua PASS nhung van nghi ngo", "PASS = khop chu ky + cau truc + logic be mat (parser heuristic). "
         "Nghiep vu quan trong van nen co test. Doc them ghi chu RULE neu co."),
    ])

    wb.save(xlsx_path)
    return xlsx_path


def main():
    rules_json = _ROOT / "rules" / "conversion_rules.json"
    rules_xlsx = _ROOT / "rules" / "conversion_rules.xlsx"
    # documents/ da chia thu muc con: markdowns/ (.md) va excels/ (.xlsx)
    guide_dir = _ROOT / "documents" / "excels"
    guide_xlsx = (guide_dir if guide_dir.is_dir()
                  else _ROOT / "documents") / "11_Tool_Usage_Guideline.xlsx"
    export_rules_excel(str(rules_json), str(rules_xlsx))
    print(f"Da xuat: {rules_xlsx}")
    export_guideline_excel(str(guide_xlsx))
    print(f"Da xuat: {guide_xlsx}")


if __name__ == "__main__":
    sys.exit(main())
